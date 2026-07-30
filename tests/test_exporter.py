"""Behavioral tests for the XLSX exporter."""

from __future__ import annotations

import importlib
import re
import tomllib
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
import pytest


def _load_exporter() -> Any:
    """Load the exporter module, allowing the RED test to assert its absence."""
    try:
        return importlib.import_module("trek_fixture.exporter")
    except ModuleNotFoundError as error:
        if error.name == "trek_fixture.exporter":
            return None
        raise


def test_openpyxl_is_a_runtime_dependency() -> None:
    """The exporter dependency is available to normal package consumers."""
    project_file = Path(__file__).parents[1] / "pyproject.toml"
    with project_file.open("rb") as file:
        project = tomllib.load(file)

    runtime_dependencies = project["project"]["dependencies"]
    is_openpyxl = re.compile(r"^openpyxl(?:\s|[<>=!~]|$)", re.IGNORECASE)

    assert any(is_openpyxl.match(dependency) for dependency in runtime_dependencies)


def test_export_xlsx_constructs_and_saves_with_openpyxl(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The exporter constructs a real workbook and saves it at the requested path."""
    exporter = _load_exporter()
    assert exporter is not None
    export_xlsx = getattr(exporter, "export_xlsx", None)
    assert callable(export_xlsx)

    real_workbook: Any = openpyxl.Workbook
    real_init = real_workbook.__init__
    real_save = real_workbook.save
    constructed_workbooks: list[Any] = []
    saved_paths: list[Path] = []

    def workbook_init(workbook: Any, *args: Any, **kwargs: Any) -> None:
        real_init(workbook, *args, **kwargs)
        constructed_workbooks.append(workbook)

    def workbook_save(
        workbook: Any, path: str | Path, *args: Any, **kwargs: Any
    ) -> Any:
        saved_paths.append(Path(path))
        return real_save(workbook, path, *args, **kwargs)

    monkeypatch.setattr(real_workbook, "__init__", workbook_init)
    monkeypatch.setattr(real_workbook, "save", workbook_save)

    output_path = tmp_path / "results.xlsx"
    export_xlsx(output_path, [("2+2", 4)])

    assert len(constructed_workbooks) == 1
    assert saved_paths == [output_path]
    assert output_path.is_file()


def test_export_xlsx_preserves_ordered_calculation_results(tmp_path: Path) -> None:
    """The saved workbook contains exactly one two-cell row for each result."""
    exporter = _load_exporter()
    assert exporter is not None
    export_xlsx = getattr(exporter, "export_xlsx", None)
    assert callable(export_xlsx)

    output_path = tmp_path / "results.xlsx"
    export_xlsx(output_path, [("2+2", 4), ("3*5", 15)])

    assert output_path.is_file()
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        assert worksheet is not None
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows == [("2+2", 4), ("3*5", 15)]
