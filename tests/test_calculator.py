"""Tests for the calculator module."""

from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from trek_fixture import calculator
from trek_fixture.calculator import add, divide, multiply, subtract


def test_add() -> None:
    assert add(2, 3) == 5
    assert add(-1, 1) == 0


def test_subtract() -> None:
    assert subtract(5, 3) == 2


def test_multiply() -> None:
    assert multiply(4, 3) == 12


def test_divide() -> None:
    assert divide(10, 2) == 5


def test_divide_by_zero_raises() -> None:
    with pytest.raises(ZeroDivisionError):
        divide(1, 0)


def test_export_xlsx_writes_results_as_ordered_rows(tmp_path: Path) -> None:
    output_path = tmp_path / "calculation-results.xlsx"
    results = [(2, 3, 5), (10, 4, 14), (-1, 8, 7)]

    calculator.export_xlsx(output_path, results)

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        assert worksheet is not None
        assert list(worksheet.iter_rows(values_only=True)) == results
    finally:
        workbook.close()

    with ZipFile(output_path) as archive:
        assert archive.testzip() is None
        assert "[Content_Types].xml" in archive.namelist()
        assert "xl/workbook.xml" in archive.namelist()
        assert "xl/worksheets/sheet1.xml" in archive.namelist()
