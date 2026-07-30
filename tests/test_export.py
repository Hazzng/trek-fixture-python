"""Tests for XLSX result export."""

from collections.abc import Callable
from decimal import Decimal
from importlib import import_module
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _get_export_xlsx() -> Callable[..., None]:
    try:
        return cast(
            Callable[..., None], import_module("trek_fixture.export").export_xlsx
        )
    except ModuleNotFoundError as exc:
        pytest.fail(f"the XLSX exporter is unavailable: {exc}")


def test_export_xlsx_round_trip(tmp_path: Path) -> None:
    """Exported results are preserved as ordered rows in a valid workbook."""
    output_path = tmp_path / "results.xlsx"
    results = [("2+2", 4), ("3*5", 15)]

    _get_export_xlsx()(output_path, results)

    assert output_path.is_file()
    workbook = load_workbook(output_path)
    try:
        assert len(workbook.worksheets) == 1
        worksheet = cast(Worksheet, workbook.active)
        assert list(worksheet.iter_rows(values_only=True)) == results
    finally:
        workbook.close()


def test_export_xlsx_preserves_formula_text_and_blank_rows(tmp_path: Path) -> None:
    """Formula-looking expressions and all-blank pairs remain visible as rows."""
    output_path = tmp_path / "special-results.xlsx"
    results = [("=1+1", 2), (None, None)]

    _get_export_xlsx()(output_path, results)

    workbook = load_workbook(output_path, data_only=False)
    try:
        worksheet = cast(Worksheet, workbook.active)
        assert worksheet["A1"].data_type == "s"
    finally:
        workbook.close()

    workbook = load_workbook(output_path, data_only=True)
    try:
        worksheet = cast(Worksheet, workbook.active)
        assert list(worksheet.iter_rows(values_only=True)) == [
            ("=1+1", 2),
            (None, None),
        ]
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "value",
    [
        float("nan"),
        float("inf"),
        float("-inf"),
        Decimal("NaN"),
        Decimal("Infinity"),
        Decimal("-Infinity"),
        Decimal("sNaN"),
    ],
)
def test_export_xlsx_preserves_non_finite_values(tmp_path: Path, value: Any) -> None:
    """Non-finite numeric values are persisted as explicit text."""
    output_path = tmp_path / "non-finite.xlsx"

    _get_export_xlsx()(output_path, [("result", value)])

    workbook = load_workbook(output_path, data_only=True)
    try:
        worksheet = cast(Worksheet, workbook.active)
        assert worksheet["B1"].value == str(value)
    finally:
        workbook.close()
