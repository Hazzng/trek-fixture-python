"""Export calculation results to an Excel workbook."""

from __future__ import annotations

from collections.abc import Iterable
from decimal import Decimal
from math import isfinite
from os import PathLike
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def _excel_value(value: Any) -> Any:
    """Return a value that openpyxl can persist without losing information."""
    if value is None:
        return ""
    if isinstance(value, float) and not isfinite(value):
        return str(value)
    if isinstance(value, Decimal) and not value.is_finite():
        return str(value)
    return value


def export_xlsx(path: str | PathLike[str], results: Iterable[tuple[Any, Any]]) -> None:
    """Save calculation expression/value pairs as ordered worksheet rows.

    Each pair in ``results`` becomes one row, with the expression in the first
    cell and its calculated value in the second cell.
    """
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    for expression, value in results:
        worksheet.append(("", _excel_value(value)))
        row = worksheet.max_row
        expression_cell = cast(Cell, worksheet.cell(row=row, column=1))
        expression_cell.value = "" if expression is None else str(expression)
        expression_cell.data_type = "s"
        value_cell = cast(Cell, worksheet.cell(row=row, column=2))
        if isinstance(value_cell.value, str):
            value_cell.data_type = "s"
    workbook.save(path)
